import io
import csv
import openpyxl
import random
import string
from typing import Optional, List, Dict
from fastapi import APIRouter, UploadFile, File, HTTPException, Form, status
from PyPDF2 import PdfReader
from app.core.database import get_supabase
from app.services.ai_parsing_service import parse_students_with_ai, parse_results_with_ai
from app.services.ml_risk_service import train_model_async
from app.services.audit_service import log_action
from app.dependencies import get_current_user
from fastapi import APIRouter, UploadFile, File, HTTPException, Form, status, Depends

BATCH_SIZE = 50   # Insert 50 students at a time


def extract_text_from_file(file_bytes: bytes, filename: str) -> str:
    """Extract all text from PDF, Excel, or CSV."""
    text = ""
    if filename.endswith(".pdf"):
        reader = PdfReader(io.BytesIO(file_bytes))
        for page in reader.pages:
            t = page.extract_text()
            if t:
                text += t + "\n"
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                text += " ".join([str(cell) for cell in row if cell is not None]) + "\n"
    elif filename.endswith(".csv"):
        try:
            text = file_bytes.decode("utf-8-sig")
        except:
            text = file_bytes.decode("latin-1")
    return text


def extract_names_from_pdf(file_bytes: bytes) -> List[dict]:
    """Extract student names from PDF. Returns list of dicts with 'name' key."""
    reader = PdfReader(io.BytesIO(file_bytes))
    names = []
    for page in reader.pages:
        text = page.extract_text()
        if not text:
            continue
        lines = text.split("\n")
        for line in lines:
            line = line.strip()
            if not line or line.isdigit():
                continue
            # If line contains a colon, treat as label: value and take the value
            if ":" in line:
                # Use only the part after the first colon, trimmed
                name = line.split(":", 1)[1].strip()
                if name:
                    names.append({"name": name, "admission_number": None, "year": None})
            else:
                # Entire line is a name
                names.append({"name": line, "admission_number": None, "year": None})
    # If nothing found, try splitting all words as names (last resort)
    if not names:
        for page in reader.pages:
            text = page.extract_text()
            if text:
                words = text.split()
                for w in words:
                    if w.isalpha() and len(w) > 1:
                        names.append({"name": w, "admission_number": None, "year": None})
    return names


def extract_names_from_excel(file_bytes: bytes) -> List[str]:
    """Read first column of Excel file (names)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    sheet = wb.active
    names = []
    for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
        if row and row[0] and isinstance(row[0], str):
            name = row[0].strip()
            if name.lower() not in ("name", "student name", "names", ""):
                names.append(name)
    return names


def extract_names_from_csv(file_bytes: bytes) -> List[str]:
    """Read first column of CSV (names)."""
    try:
        text = file_bytes.decode("utf-8-sig")
    except:
        text = file_bytes.decode("latin-1")
    reader = csv.reader(io.StringIO(text))
    names = []
    for row in reader:
        if row and row[0].strip():
            name = row[0].strip()
            if name.lower() not in ("name", "student name", "names", ""):
                names.append(name)
    return names


@router.post("/students")
async def import_students(
    school_id: str = Form(...),
    class_id: str = Form(...),
    file: UploadFile = File(...),
    use_ai: bool = Form(False),
    current_user: dict = Depends(get_current_user)
):
    if school_id != current_user["school_id"]:
        raise HTTPException(status_code=403, detail="Forbidden")
        
    db = get_supabase()

    # Validate school and class
    school = db.table("schools").select("id").eq("id", school_id).execute()
    if not school.data:
        raise HTTPException(status_code=404, detail="School not found")
    cls = db.table("classes").select("id").eq("id", class_id).eq("school_id", school_id).execute()
    if not cls.data:
        raise HTTPException(status_code=404, detail="Class not found or not in this school")

    file_bytes = await file.read()
    filename = file.filename.lower() if file.filename else ""

    # Extract raw names (list of dicts with at least "name")
    raw_students = []
    
    if use_ai:
        text = extract_text_from_file(file_bytes, filename)
        raw_students = await parse_students_with_ai(text)
    
    if not raw_students:
        if filename.endswith(".pdf"):
            raw_students = extract_names_from_pdf(file_bytes)
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            names = extract_names_from_excel(file_bytes)
            raw_students = [{"name": n} for n in names]
        elif filename.endswith(".csv"):
            names = extract_names_from_csv(file_bytes)
            raw_students = [{"name": n} for n in names]
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type. Use PDF, Excel (.xlsx) or CSV.")

    if not raw_students:
        raise HTTPException(status_code=400, detail="No names found in file. Check the format or the file may be empty.")

    # Get the current maximum admission number for this school to generate new ones
    prefix = school_id[:4].upper()
    existing = db.table("students").select("admission_number").eq("school_id", school_id).order("admission_number", desc=True).limit(100).execute()
    max_num = 0
    if existing.data:
        for row in existing.data:
            last_adm = row.get("admission_number")
            if not last_adm: continue
            try:
                # Extract numeric part (last digits)
                num_part = ""
                for char in reversed(str(last_adm)):
                    if char.isdigit():
                        num_part = char + num_part
                    else:
                        break
                if num_part:
                    max_num = max(max_num, int(num_part))
                    break # Found the highest one
            except:
                continue

    # Prepare data for batch insert
    to_insert = []
    errors = []
    for idx, student in enumerate(raw_students):
        name = student.get("name")
        if not name: continue
        
        admission_num = student.get("admission_number")
        if not admission_num:
            max_num += 1
            admission_num = f"{prefix}{max_num:04d}"
            
        access_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
        to_insert.append({
            "school_id": school_id,
            "class_id": class_id,
            "name": name,
            "admission_number": admission_num,
            "access_code": access_code,
        })

    # Batch insert
    inserted = 0
    for i in range(0, len(to_insert), BATCH_SIZE):
        batch = to_insert[i:i + BATCH_SIZE]
        try:
            result = db.table("students").insert(batch).execute()
            inserted += len(result.data) if result.data else 0
        except Exception as e:
            for student in batch:
                try:
                    db.table("students").insert(student).execute()
                    inserted += 1
                except Exception as single_err:
                    errors.append(f"{student['name']}: {str(single_err)}")

    if inserted > 0:
        # Update student count (manual loop since we don't have bulk increment yet)
        # or just set it to the count
        db.table("schools").update({
            "student_count": db.table("students").select("id", count="exact").eq("school_id", school_id).execute().count
        }).eq("id", school_id).execute()

        log_action(
            school_id=school_id,
            action="STUDENTS_BULK_IMPORTED",
            actor_id=current_user["id"],
            actor_name=current_user["name"],
            entity_type="school",
            entity_id=school_id,
            new_value={"count": inserted}
        )

    return {"message": f"{inserted} students imported successfully", "count": inserted, "errors": errors}


@router.post("/results")
async def import_results(
    school_id: str = Form(...),
    teacher_id: str = Form(...),
    class_id: Optional[str] = Form(None),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    if school_id != current_user["school_id"]:
         raise HTTPException(status_code=403, detail="Forbidden")

    db = get_supabase()

    # 1. Verify school and teacher
    teacher = db.table("teachers").select("role").eq("id", teacher_id).eq("school_id", school_id).execute()
    if not teacher.data:
        raise HTTPException(status_code=404, detail="Teacher not found in this school")
    
    # 2. Extract text and parse with AI
    file_bytes = await file.read()
    filename = file.filename.lower() if file.filename else ""
    text = extract_text_from_file(file_bytes, filename)
    
    if not text.strip():
        raise HTTPException(status_code=400, detail="Could not extract text from file")
        
    ai_results = await parse_results_with_ai(text)
    if not ai_results:
        raise HTTPException(status_code=400, detail="AI could not identify any results in the document")

    # 3. Fetch school data for mapping
    students_db = db.table("students").select("id, name, admission_number, class_id").eq("school_id", school_id).execute().data or []
    subjects_db = db.table("subjects").select("id, name").execute().data or []
    
    # Pre-map for performance
    adm_to_student = {s["admission_number"]: s for s in students_db}
    name_to_student = {s["name"].lower(): s for s in students_db}
    name_to_subject = {s["name"].lower(): s["id"] for s in subjects_db}

    # 4. Process AI output
    to_insert = []
    errors = []
    
    for item in ai_results:
        st_id = None
        matched_student = None
        # Match student
        if item.get("admission_number") and item["admission_number"] in adm_to_student:
            matched_student = adm_to_student[item["admission_number"]]
            st_id = matched_student["id"]
        elif item.get("name") and item["name"].lower() in name_to_student:
            matched_student = name_to_student[item["name"].lower()]
            st_id = matched_student["id"]
        
        if not st_id:
            errors.append(f"Student not found: {item.get('name') or item.get('admission_number')}")
            continue

        # Match subject
        sub_id = None
        if item.get("subject"):
            sub_name = item["subject"].lower()
            sub_id = name_to_subject.get(sub_name)
            if not sub_id:
                # Fuzzy match
                for name, sid in name_to_subject.items():
                    if sub_name in name or name in sub_name:
                        sub_id = sid
                        break
        
        if not sub_id:
            errors.append(f"Subject not found: {item.get('subject')}")
            continue

        # Build row
        try:
            score_val = float(item.get("score") or 0)
        except:
            score_val = 0

        to_insert.append({
            "student_id": st_id,
            "subject_id": sub_id,
            "class_id": matched_student["class_id"],
            "exam_type": item.get("exam_type", "EXAM") if item.get("exam_type") in ("CAT", "EXAM") else "EXAM",
            "term": item.get("term") or "Unknown Term",
            "academic_year": str(item.get("academic_year") or "2025"),
            "score": score_val,
            "submitted_by": teacher_id
        })

    # 5. Batch insert
    inserted = 0
    if to_insert:
        for i in range(0, len(to_insert), BATCH_SIZE):
            chunk = to_insert[i:i + BATCH_SIZE]
            try:
                res = db.table("results").insert(chunk).execute()
                inserted += len(res.data) if res.data else 0
            except Exception as e:
                errors.append(f"Batch failed: {str(e)}")

    # Trigger ML training if enough new data was added
    if inserted > 0:
        train_model_async()

    return {
        "message": f"Successfully imported {inserted} results",
        "count": inserted,
        "errors": errors,
        "raw_ai_count": len(ai_results)
    }
